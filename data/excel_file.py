import os
from typing import Optional

import openpyxl

from data.excel_sheet import ExcelSheet


class ExcelFile:
    def __init__(self, filename: str, create_file=False):
        self.__filename = filename
        self.__excel_sheet_list = []
        if create_file:
            self.__workbook = openpyxl.Workbook()
        else:
            self.__workbook = openpyxl.load_workbook(filename)

        for worksheet in self.__workbook.worksheets:
            self.__excel_sheet_list.append(ExcelSheet(worksheet))

        self.__tmp_active = None
        self.__tmp_active_sheet = None

    def __get_active_sheet(self) -> Optional[ExcelSheet]:
        if self.__workbook.active is None:
            return None

        if self.__workbook.active is self.__tmp_active:
            return self.__tmp_active_sheet

        self.__tmp_active = self.__workbook.active
        self.__tmp_active_sheet = ExcelSheet(self.__workbook.active)
        return self.__tmp_active_sheet

    active_sheet = property(__get_active_sheet)

    def is_contains(self, sheet_title) -> bool:
        is_contains = False
        for excel_sheet in self.__excel_sheet_list:
            if sheet_title == excel_sheet.get_title():
                is_contains = True

        return is_contains

    def get_sheet_from_index(self, index: int):
        if len(self.__excel_sheet_list) > index:
            return self.__excel_sheet_list[index]

    def get_sheet_from_sheet_title(self, sheet_title: str) -> ExcelSheet:
        for excel_sheet in self.__excel_sheet_list:
            if sheet_title == excel_sheet.get_title():
                return excel_sheet

    def create_sheet(self, sheet_title: str, index=None):

        if index is None:
            worksheet = self.__workbook.create_sheet(title=sheet_title)
            excel_sheet = ExcelSheet(worksheet)
            self.__excel_sheet_list.append(excel_sheet)
            self.__workbook.active = self.__workbook.sheetnames.index(sheet_title)
        else:
            worksheet = self.__workbook.create_sheet(title=sheet_title, index=index)
            excel_sheet = ExcelSheet(worksheet)
            self.__excel_sheet_list.insert(index, excel_sheet)
            self.__workbook.active = index

    def delete_sheet_from_index(self, index: int):
        if len(self.__excel_sheet_list) > index:
            self.__workbook.remove(self.__workbook.worksheets[index])
            excel_sheet = self.get_sheet_from_index(index)
            self.__excel_sheet_list.remove(excel_sheet)

    def delete_sheet_from_sheet_title(self, sheet_title: str):
        if self.is_contains(sheet_title):
            self.__workbook.remove(self.__workbook[sheet_title])
            excel_sheet = self.get_sheet_from_sheet_title(sheet_title)
            self.__excel_sheet_list.remove(excel_sheet)

    def get_sheet_count(self) -> int:
        return len(self.__excel_sheet_list)

    def switch_active_sheet_from_index(self, index: int):
        if len(self.__excel_sheet_list) > index:
            self.__workbook.active = index

    def switch_active_sheet_from_sheet_title(self, sheet_title: str):
        if self.is_contains(sheet_title):
            self.__workbook.active = self.__workbook.sheetnames.index(sheet_title)

    def save(self, filename=None):
        if filename is None:
            self.__workbook.save(self.__filename)
        else:
            self.__workbook.save(filename)

    def delete(self):
        try:
            os.remove(self.__filename)
        except IOError:
            pass

    def __str__(self):
        description = f'{self.__filename} \n'
        description += '\n'.join(
            [f'{index} : {sheetname}' for index, sheetname in enumerate(self.__workbook.sheetnames)])
        return description


class ExcelFileFactory:
    @staticmethod
    def load(filename):
        return ExcelFile(filename=filename, create_file=False)

    @staticmethod
    def create(filename):
        return ExcelFile(filename=filename, create_file=True)
