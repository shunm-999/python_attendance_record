from collections import Iterable

from openpyxl.styles import Font, Side, Border, PatternFill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation

from const.alignmentconst import AlignmentConst
from util.excel_util import ExcelUtil


class ExcelSheet:
    MAX_ROW = 1048576

    def __init__(self, worksheet):
        self.__worksheet = worksheet

    def get_title(self) -> str:
        return self.__worksheet.title

    def get(self, row, column):
        if type(column) is str:
            column = column_index_from_string(column)
        return self.__worksheet.cell(row=row, column=column).value

    def get_rows(self, start_row, start_column, end_row=None, end_column=None):
        cell_range = ExcelUtil.create_cell_range_string(
            start_row=start_row,
            start_column=start_column,
            end_row=end_row,
            end_column=end_column)

        values = []
        for row in self.__worksheet[cell_range]:
            row_value = []
            for col in row:
                row_value.append(col.value)
            values.append(row_value)

        return values

    def put(self, row, column, value):
        if type(column) is str:
            column = column_index_from_string(column)
        self.__worksheet.cell(row=row, column=column).value = value

    def put_rows(self, start_row, start_column, end_row=None, end_column=None, values=None):
        if values is None:
            values = [[]]
        if not isinstance(values, Iterable):
            return

        if not isinstance(values[0], Iterable):
            return

        cell_range = ExcelUtil.create_cell_range_string(
            start_row=start_row,
            start_column=start_column,
            end_row=end_row,
            end_column=end_column)

        for row_num, row in enumerate(self.__worksheet[cell_range]):
            for column_num, col in enumerate(row):
                col.value = values[row_num][column_num]

    def insert_rows(self, row):
        self.__worksheet.insert_rows(row)

    def delete_rows(self, row):
        self.__worksheet.delete_rows(row)

    def insert_columns(self, column):
        if type(column) is str:
            column = column_index_from_string(column)
        self.__worksheet.insert_cols(column)

    def delete_columns(self, column):
        if type(column) is str:
            column = column_index_from_string(column)
        self.__worksheet.delete_cols(column)

    def get_row_height(self, row) -> int:
        return self.__worksheet.row_dimensions[row].height

    def set_row_height(self, row, height):
        self.__worksheet.row_dimensions[row].height = height

    def get_column_width(self, column) -> int:
        if type(column) is int:
            column = get_column_letter(column)
        return self.__worksheet.column_dimensions[column].width

    def set_column_width(self, column, width):
        if type(column) is int:
            column = get_column_letter(column)
        self.__worksheet.column_dimensions[column].width = width

    def merge_cells(self, start_row, start_column, end_row, end_column):

        if type(start_column) is str:
            start_column = column_index_from_string(start_column)
        if type(end_column) is str:
            end_column = column_index_from_string(end_column)

        merge_cell_range = ExcelUtil.create_cell_range_string(
            start_row=start_row,
            start_column=start_column,
            end_row=end_row,
            end_column=end_column)

        self.__worksheet.merge_cells(merge_cell_range)

    def set_freeze_panes(self, row, column):
        if type(column) is int:
            column = get_column_letter(column)

        self.__worksheet.freeze_panes = f'{column}{row}'

    def get_font(self, row, column) -> Font:
        if type(column) is str:
            column = column_index_from_string(column)
        return self.__worksheet.cell(row=row, column=column).font

    def set_font(self, row, column, size=None, bold=None, italic=None):
        if type(column) is str:
            column = column_index_from_string(column)
        self.__worksheet.cell(row=row, column=column).font = Font(size=size, bold=bold, italic=italic)

    def get_border(self, row, column) -> Border:
        if type(column) is str:
            column = column_index_from_string(column)
        return self.__worksheet.cell(row=row, column=column).border

    def set_border(self, row, column, top=False, bottom=False, left=False, right=False, style='thin',
                   color='000000') -> None:
        if type(column) is str:
            column = column_index_from_string(column)
        side = Side(style=style, color=color)
        top = side if top else None
        bottom = side if bottom else None
        left = side if left else None
        right = side if right else None

        border = Border(top=top, bottom=bottom, left=left, right=right)

        self.__worksheet.cell(row=row, column=column).border = border

    def set_ruled_line(self, start_row, start_column, end_row=None, end_column=None, style='thin', color='000000'):

        if end_row is None:
            end_row = start_row
        if end_column is None:
            end_column = end_column

        side = Side(style=style, color=color)
        border = Border(top=side, bottom=side, left=side, right=side)

        if type(start_column) is str:
            start_column = column_index_from_string(start_column)
        if type(end_column) is str:
            end_column = column_index_from_string(end_column)

        for row_num in range(start_row, end_row + 1):
            for column_num in range(start_column, end_column + 1):
                self.__worksheet.cell(row=row_num, column=column_num).border = border

    def get_pattern_fill(self, row, column) -> PatternFill:
        if type(column) is str:
            column = column_index_from_string(column)

        return self.__worksheet.cell(row=row, column=column).fill

    def set_pattern_fill(self,
                         start_row,
                         start_column,
                         end_row=None,
                         end_column=None,
                         pattern_type='solid',
                         fg_color='d3d3d3'):
        if end_row is None:
            end_row = start_row
        if end_column is None:
            end_column = start_column

        if type(start_column) is str:
            start_column = column_index_from_string(start_column)
        if type(end_column) is str:
            end_column = column_index_from_string(end_column)

        fill = PatternFill(patternType=pattern_type, fgColor=fg_color)

        for row_num in range(start_row, end_row + 1):
            for column_num in range(start_column, end_column + 1):
                self.__worksheet.cell(row=row_num, column=column_num).fill = fill

    def set_data_validation(self, start_row, end_row, column, formula=None):
        if formula is None:
            formula = []

        if end_row is None:
            end_row = ExcelSheet.MAX_ROW

        formula = ','.join(formula)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)

        dv.ranges = ExcelUtil.create_cell_range_string(
            start_row=start_row,
            start_column=column,
            end_row=end_row,
            end_column=column)

        self.__worksheet.add_data_validation(dv)

    def set_data_validation_from_cell_range(self, start_row, end_row, column, cell_range: str):

        if end_row is None:
            end_row = ExcelSheet.MAX_ROW

        dv = DataValidation(type="list", formula1=cell_range, allow_blank=True)

        dv.ranges = ExcelUtil.create_cell_range_string(
            start_row=start_row,
            start_column=column,
            end_row=end_row,
            end_column=column)

        self.__worksheet.add_data_validation(dv)

    def set_number_format(self, row, column, number_format: str):
        coordinate = ExcelUtil.create_cell_string(row=row, column=column)
        self.__worksheet[coordinate].number_format = number_format

    def set_number_formats(self, start_row, start_column, end_row, end_column, number_format: str):
        if type(start_column) is str:
            start_column = column_index_from_string(start_column)
        if type(end_column) is str:
            end_column = column_index_from_string(end_column)

        for row_num in range(start_row, end_row + 1):
            for column_num in range(start_column, end_column + 1):
                self.__worksheet.cell(row=row_num, column=column_num).number_format = number_format

    def set_text_alignment(self, row, column, horizontal: AlignmentConst, vertical: AlignmentConst,
                           text_rotation: int = 0):
        coordinate = ExcelUtil.create_cell_string(row=row, column=column)
        self.__worksheet[coordinate].alignment = Alignment(horizontal=horizontal.value, vertical=vertical.value,
                                                           textRotation=text_rotation)

    def set_text_alignments(self, start_row, start_column, end_row, end_column, horizontal: AlignmentConst,
                            vertical: AlignmentConst, text_rotation: int = 0):
        if type(start_column) is str:
            start_column = column_index_from_string(start_column)
        if type(end_column) is str:
            end_column = column_index_from_string(end_column)

        for row_num in range(start_row, end_row + 1):
            for column_num in range(start_column, end_column + 1):
                self.__worksheet.cell(row=row_num, column=column_num).alignment = Alignment(horizontal=horizontal.value,
                                                                                            vertical=vertical.value,
                                                                                            textRotation=text_rotation)
