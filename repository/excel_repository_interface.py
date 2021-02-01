from abc import ABCMeta, abstractmethod

from openpyxl.styles import Border, Font, PatternFill

from const.alignmentconst import AlignmentConst
from data.excel_sheet import ExcelSheet


class ExcelRepositoryInterface(metaclass=ABCMeta):
    @abstractmethod
    def is_contains(self, sheet_title) -> bool:
        pass

    @abstractmethod
    def get_sheet_from_index(self, index: int):
        pass

    @abstractmethod
    def get_sheet_from_sheet_title(self, sheet_title: str) -> ExcelSheet:
        pass

    @abstractmethod
    def get_active_sheet(self) -> ExcelSheet:
        pass

    @abstractmethod
    def create_sheet(self, sheet_title: str, index=None):
        pass

    @abstractmethod
    def delete_sheet_from_index(self, index: int):
        pass

    @abstractmethod
    def delete_sheet_from_sheet_title(self, sheet_title: str):
        pass

    @abstractmethod
    def get_sheet_count(self) -> int:
        pass

    @abstractmethod
    def switch_active_sheet_from_index(self, index: int):
        pass

    @abstractmethod
    def switch_active_sheet_from_sheet_title(self, sheet_title: str):
        pass

    @abstractmethod
    def save(self, filename):
        pass

    """
    基本関数
    """

    @abstractmethod
    def get(self, row, column):
        pass

    @abstractmethod
    def get_rows(self, start_row, start_column, end_row=None, end_column=None):
        pass

    @abstractmethod
    def put(self, row, column, value):
        pass

    @abstractmethod
    def put_rows(self, start_row, start_column, end_row=None, end_column=None, values=None):
        pass

    @abstractmethod
    def insert_rows(self, row):
        pass

    @abstractmethod
    def delete_rows(self, row):
        pass

    @abstractmethod
    def insert_columns(self, column):
        pass

    @abstractmethod
    def delete_columns(self, column):
        pass

    @abstractmethod
    def get_row_height(self, row) -> int:
        pass

    @abstractmethod
    def set_row_height(self, row, height):
        pass

    @abstractmethod
    def get_column_width(self, column) -> int:
        pass

    @abstractmethod
    def set_column_width(self, column, width):
        pass

    @abstractmethod
    def merge_cells(self, start_row, start_column, end_row, end_column):
        pass

    @abstractmethod
    def set_freeze_panes(self, row, column):
        pass

    @abstractmethod
    def get_font(self, row, column) -> Font:
        pass

    @abstractmethod
    def set_font(self, row, column, size=None, bold=None, italic=None):
        pass

    @abstractmethod
    def get_border(self, row, column) -> Border:
        pass

    @abstractmethod
    def set_border(self, row, column, top=False, bottom=False, left=False, right=False, style='thin',
                   color='000000') -> None:
        pass

    @abstractmethod
    def set_ruled_line(self, start_row, start_column, end_row=None, end_column=None, style='thin', color='000000'):
        pass

    @abstractmethod
    def get_pattern_fill(self, row, column) -> PatternFill:
        pass

    @abstractmethod
    def set_pattern_fill(self,
                         start_row,
                         start_column,
                         end_row=None,
                         end_column=None,
                         pattern_type='solid',
                         fg_color='d3d3d3'):
        pass

    @abstractmethod
    def set_data_validation(self, start_row, end_row, column, formula=None):
        pass

    @abstractmethod
    def set_data_validation_from_cell_range(self, start_row, end_row, column, cell_range: str):
        pass

    @abstractmethod
    def set_number_format(self, row, column, number_format: str):
        pass

    @abstractmethod
    def set_number_formats(self, start_row, start_column, end_row, end_column, number_format: str):
        pass

    @abstractmethod
    def set_text_alignment(self, row, column, horizontal: AlignmentConst, vertical: AlignmentConst,
                           text_rotation: int = 0):
        pass

    @abstractmethod
    def set_text_alignments(self, start_row, start_column, end_row, end_column, horizontal: AlignmentConst,
                            vertical: AlignmentConst, text_rotation: int = 0):
        pass

    """
    Excel関数
    """

    @abstractmethod
    def put_sum(self, target_row, target_column, start_row, start_column, end_row, end_column):
        pass

    @abstractmethod
    def put_average(self, target_row, target_column, start_row, start_column, end_row, end_column):
        pass
