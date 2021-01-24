import os
import unittest

from openpyxl.styles import Color
from openpyxl.styles.colors import BLACK

from data.excel_sheet import ExcelSheet
from repository.excel_repository import ExcelRepository
from repository.excel_repository_interface import ExcelRepositoryInterface
from util.excel_util import ExcelUtil


class TestEmptyExcelRepository(unittest.TestCase):

    def setUp(self) -> None:
        self.excel_filename: str = 'sample.xlsx'
        self.excel_repository: ExcelRepositoryInterface = ExcelRepository(self.excel_filename)
        self.excel_repository.delete_sheet_from_index(0)

    def tearDown(self) -> None:
        self.excel_repository.delete()

    def test_init_repository(self):
        self.excel_repository.create_sheet('TestSheet1')
        active_sheet: ExcelSheet = self.excel_repository.get_active_sheet()
        self.assertEqual('TestSheet1', active_sheet.get_title())

    """
    ファイル操作
    """

    def test_is_contains(self):
        self.excel_repository.create_sheet('sample')
        self.assertTrue(self.excel_repository.is_contains('sample'))

    def test_get_sheet_from_index(self):
        expected = 'sample'
        self.excel_repository.create_sheet(expected)

        actual = self.excel_repository.get_sheet_from_index(0).get_title()
        self.assertEqual(expected, actual)

    def test_get_sheet_from_index_none(self):
        self.assertIsNone(self.excel_repository.get_sheet_from_index(0))

    def test_get_sheet_from_sheet_title(self):
        expected = 'sample'
        self.excel_repository.create_sheet(expected)

        actual = self.excel_repository.get_sheet_from_sheet_title(expected).get_title()
        self.assertEqual(expected, actual)

    def test_get_sheet_from_sheet_title_none(self):
        self.assertIsNone(self.excel_repository.get_sheet_from_sheet_title('sample'))

    def test_get_active_sheet(self):
        expected = 'sample'
        self.excel_repository.create_sheet(expected)

        actual = self.excel_repository.get_active_sheet().get_title()
        self.assertEqual(expected, actual)

    def test_get_active_sheet_none(self):
        self.assertIsNone(self.excel_repository.get_active_sheet())

    def test_create_sheet(self):
        expected_sheet1 = 'sample'
        expected_sheet2 = 'sample2'

        self.excel_repository.create_sheet(expected_sheet1)
        actual_sheet1 = self.excel_repository.get_sheet_from_sheet_title(expected_sheet1).get_title()
        self.assertEqual(expected_sheet1, actual_sheet1)

        self.excel_repository.create_sheet(expected_sheet2)
        actual_sheet2 = self.excel_repository.get_sheet_from_sheet_title(expected_sheet2).get_title()
        self.assertEqual(expected_sheet2, actual_sheet2)

    def test_insert_sheet(self):
        expected_sheet1 = 'sample'
        expected_sheet2 = 'sample2'

        self.excel_repository.create_sheet(expected_sheet1, index=0)
        actual_sheet1 = self.excel_repository.get_sheet_from_index(0).get_title()
        self.assertEqual(expected_sheet1, actual_sheet1)

        self.excel_repository.create_sheet(expected_sheet2, index=0)
        actual_sheet2 = self.excel_repository.get_sheet_from_index(0).get_title()
        self.assertEqual(expected_sheet2, actual_sheet2)

    def test_delete_sheet_from_index(self):
        self.excel_repository.create_sheet('sample')
        self.assertEqual(1, self.excel_repository.get_sheet_count())

        self.excel_repository.delete_sheet_from_index(0)
        self.assertEqual(0, self.excel_repository.get_sheet_count())

    def test_delete_sheet_from_wrong_index(self):
        self.excel_repository.create_sheet('sample')
        self.assertEqual(1, self.excel_repository.get_sheet_count())

        self.excel_repository.delete_sheet_from_index(2)
        self.assertEqual(1, self.excel_repository.get_sheet_count())

    def test_delete_current_sheet_from_index(self):
        expected_sheet1 = 'sample'
        expected_sheet2 = 'sample2'

        self.excel_repository.create_sheet(expected_sheet1)
        self.excel_repository.create_sheet(expected_sheet2)

        self.assertEqual(expected_sheet2, self.excel_repository.get_active_sheet().get_title())

        self.excel_repository.delete_sheet_from_index(1)
        self.assertEqual(1, self.excel_repository.get_sheet_count())
        self.assertIsNone(self.excel_repository.get_active_sheet())

        self.excel_repository.switch_active_sheet_from_index(0)
        self.assertEqual(expected_sheet1, self.excel_repository.get_active_sheet().get_title())

    def test_delete_sheet_from_sheet_title(self):
        self.excel_repository.create_sheet('sample')
        self.assertEqual(1, self.excel_repository.get_sheet_count())

        self.excel_repository.delete_sheet_from_sheet_title('sample')
        self.assertEqual(0, self.excel_repository.get_sheet_count())

    def test_delete_sheet_from_wrong_sheet_title(self):
        self.excel_repository.create_sheet('sample')
        self.assertEqual(1, self.excel_repository.get_sheet_count())

        self.excel_repository.delete_sheet_from_sheet_title('wrong')
        self.assertEqual(1, self.excel_repository.get_sheet_count())

    def test_delete_current_sheet_from_title(self):
        expected_sheet1 = 'sample'
        expected_sheet2 = 'sample2'

        self.excel_repository.create_sheet(expected_sheet1)
        self.excel_repository.create_sheet(expected_sheet2)

        self.assertEqual(expected_sheet2, self.excel_repository.get_active_sheet().get_title())

        self.excel_repository.delete_sheet_from_sheet_title(expected_sheet2)
        self.assertEqual(1, self.excel_repository.get_sheet_count())
        self.assertIsNone(self.excel_repository.get_active_sheet())

        self.excel_repository.switch_active_sheet_from_sheet_title(expected_sheet1)
        self.assertEqual(expected_sheet1, self.excel_repository.get_active_sheet().get_title())

    def test_get_sheet_count(self):
        self.excel_repository.create_sheet('sample')
        self.excel_repository.create_sheet('sample2')

        self.assertEqual(2, self.excel_repository.get_sheet_count())

    def test_switch_active_sheet_from_index(self):
        expected1 = 'sample'
        expected2 = 'sample2'

        self.excel_repository.create_sheet(expected1)
        self.excel_repository.create_sheet(expected2)
        excel_sheet = self.excel_repository.get_active_sheet()
        self.assertEqual(expected2, excel_sheet.get_title())

        self.excel_repository.switch_active_sheet_from_index(0)
        excel_sheet = self.excel_repository.get_active_sheet()
        self.assertEqual(expected1, excel_sheet.get_title())

    def test_switch_active_sheet_from_wrong_index(self):
        expected1 = 'sample'
        expected2 = 'sample2'

        self.excel_repository.create_sheet(expected1)
        self.excel_repository.create_sheet(expected2)
        excel_sheet = self.excel_repository.get_active_sheet()
        self.assertEqual(expected2, excel_sheet.get_title())

        self.excel_repository.switch_active_sheet_from_index(2)
        excel_sheet = self.excel_repository.get_active_sheet()
        self.assertEqual(expected2, excel_sheet.get_title())

    def test_switch_active_sheet_from_sheet_title(self):
        expected1 = 'sample'
        expected2 = 'sample2'

        self.excel_repository.create_sheet(expected1)
        self.excel_repository.create_sheet(expected2)
        excel_sheet = self.excel_repository.get_active_sheet()
        self.assertEqual(expected2, excel_sheet.get_title())

        self.excel_repository.switch_active_sheet_from_sheet_title(expected1)
        excel_sheet = self.excel_repository.get_active_sheet()
        self.assertEqual(expected1, excel_sheet.get_title())

    def test_switch_active_sheet_from_wrong_sheet_title(self):
        expected1 = 'sample'
        expected2 = 'sample2'

        self.excel_repository.create_sheet(expected1)
        self.excel_repository.create_sheet(expected2)
        excel_sheet = self.excel_repository.get_active_sheet()
        self.assertEqual(expected2, excel_sheet.get_title())

        self.excel_repository.switch_active_sheet_from_sheet_title('wrong')
        excel_sheet = self.excel_repository.get_active_sheet()
        self.assertEqual(expected2, excel_sheet.get_title())

    def test_save(self):
        self.excel_repository.create_sheet('sample')
        self.assertFalse(os.path.exists(self.excel_filename))

        self.excel_repository.save()
        self.assertTrue(os.path.exists(self.excel_filename))

    def test_save_new_title(self):
        self.excel_repository.create_sheet('sample')
        self.assertFalse(os.path.exists(self.excel_filename))

        self.excel_repository.save(filename=self.excel_filename)
        self.assertTrue(os.path.exists(self.excel_filename))

    def test_delete(self):
        self.excel_repository.create_sheet('sample')
        self.excel_repository.save()

        self.assertTrue(os.path.exists(self.excel_filename))

        self.excel_repository.delete()
        self.assertFalse(os.path.exists(self.excel_filename))

    def test_delete_not_exist_file(self):
        self.excel_repository.delete()


class TestOneSheetExcelRepository(unittest.TestCase):

    def setUp(self) -> None:
        self.excel_filename: str = 'sample.xlsx'
        self.excel_repository: ExcelRepositoryInterface = ExcelRepository(self.excel_filename)
        self.excel_repository.create_sheet('TestSheet')

    """
    シート操作
    """

    def test_get_empty_cell(self):
        expected = None
        actual = self.excel_repository.get(row=1, column=1)
        self.assertEqual(expected, actual)

    def test_get_cell_from_column_number(self):
        expected = 1
        self.excel_repository.put(row=1, column=1, value=expected)

        actual = self.excel_repository.get(row=1, column=1)
        self.assertEqual(expected, actual)

    def test_get_cell_from_column_letter(self):
        expected = 1
        self.excel_repository.put(row=1, column='A', value=expected)

        actual = self.excel_repository.get(row=1, column='A')
        self.assertEqual(expected, actual)

    def test_set_and_get_rows(self):
        expected = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]

        for row_num in range(1, 4):
            for column_num in range(1, 4):
                self.excel_repository.put(row=row_num, column=column_num, value=expected[row_num - 1][column_num - 1])

        actual = self.excel_repository.get_rows(start_row=1, start_column=1, end_row=3, end_column=3)
        self.assertEqual(expected, actual)

    def test_insert_rows(self):
        expected = 1
        self.excel_repository.put(row=1, column=1, value=expected)
        self.excel_repository.insert_rows(row=1)

        actual = self.excel_repository.get(row=2, column=1)
        self.assertEqual(expected, actual)

    def test_delete_rows(self):
        expected = 1
        self.excel_repository.put(row=2, column=1, value=expected)
        self.excel_repository.delete_rows(row=1)

        actual = self.excel_repository.get(row=1, column=1)
        self.assertEqual(expected, actual)

    def test_insert_columns(self):
        expected = 1
        self.excel_repository.put(row=1, column=1, value=expected)
        self.excel_repository.insert_columns(column=1)

        actual = self.excel_repository.get(row=1, column=2)
        self.assertEqual(expected, actual)

    def test_delete_columns(self):
        expected = 1
        self.excel_repository.put(row=1, column=2, value=expected)
        self.excel_repository.delete_columns(column=1)

        actual = self.excel_repository.get(row=1, column=1)
        self.assertEqual(expected, actual)

    def test_set_and_get_row_height(self):
        expected = 16
        self.excel_repository.set_row_height(row=1, height=expected)
        actual = self.excel_repository.get_row_height(row=1)

        self.assertEqual(expected, actual)

    def test_set_and_get_column_width(self):
        expected = 16
        self.excel_repository.set_column_width(column=1, width=expected)
        actual = self.excel_repository.get_column_width(column=1)

        self.assertEqual(expected, actual)

    def test_merge_cells(self):

        self.excel_repository.put(row=1, column=1, value=1)
        self.excel_repository.put(row=3, column=3, value=3)
        self.excel_repository.merge_cells(start_row=1, start_column=1, end_row=3, end_column=3)

        expected_a1 = 1

        actual_a1 = self.excel_repository.get(row=1, column=1)
        actual_c3 = self.excel_repository.get(row=3, column=3)

        self.assertEqual(expected_a1, actual_a1)
        self.assertIsNone(actual_c3)

    def test_merge_cells_and_dimensions(self):

        for row_num in range(1, 4):
            self.excel_repository.set_row_height(row=row_num, height=16)
        for column_num in range(1, 4):
            self.excel_repository.set_column_width(column=column_num, width=16)

        self.excel_repository.merge_cells_and_dimensions(start_row=1, start_column=1, end_row=3, end_column=3)

        expected = 48

        actual_height = self.excel_repository.get_row_height(row=1)
        actual_width = self.excel_repository.get_column_width(column=1)

        self.assertEqual(expected, actual_height)
        self.assertEqual(expected, actual_width)

    def test_freeze_panes(self):
        self.excel_repository.set_freeze_panes(row=2, column=1)

    def test_freeze_panes_column_letter(self):
        self.excel_repository.set_freeze_panes(row=2, column='B')

    def test_set_and_get_font(self):
        self.excel_repository.set_font(row=1, column=1, size=16, bold=True, italic=True)
        font = self.excel_repository.get_font(row=1, column=1)

        self.assertEqual(font.size, 16)
        self.assertTrue(font.bold)
        self.assertTrue(font.italic)

    def test_set_and_get_font_column_letter(self):
        self.excel_repository.set_font(row=1, column='A', size=16, bold=True, italic=True)
        font = self.excel_repository.get_font(row=1, column='A')

        self.assertEqual(font.size, 16)
        self.assertTrue(font.bold)
        self.assertTrue(font.italic)

    def test_set_ruled_line(self):
        self.excel_repository.set_ruled_line(start_row=1, start_column=1, end_row=3, end_column=3)

        for row_num in range(1, 4):
            for column_num in range(1, 4):
                border = self.excel_repository.get_border(row=row_num, column=column_num)
                self.assertIsNotNone(border)

    def test_set_ruled_line_raise_exception(self):
        with self.assertRaises(expected_exception=ValueError):
            self.excel_repository.set_ruled_line(start_row=1, start_column=1, end_row=3, end_column=3, style='aaa')

    def test_set_and_get_pattern_fill(self):

        expected_pattern_type = 'solid'
        expected_fg_color = '00d3d3d3'

        self.excel_repository.set_pattern_fill(start_row=1, start_column=1, end_row=1, end_column=1,
                                               pattern_type=expected_pattern_type,
                                               fg_color=expected_fg_color)

        pattern_fill = self.excel_repository.get_pattern_fill(row=1, column=1)
        self.assertEqual(expected_pattern_type, pattern_fill.patternType)
        self.assertEqual(expected_fg_color, pattern_fill.fgColor.value)

    def test_set_and_get_color_from_name(self):

        expected_fg_color = Color(rgb=BLACK)
        self.excel_repository.set_pattern_fill(start_row=1, start_column=1, end_row=1, end_column=1,
                                               pattern_type='solid',
                                               fg_color=expected_fg_color)

        pattern_fill = self.excel_repository.get_pattern_fill(row=1, column=1)
        self.assertEqual(expected_fg_color, pattern_fill.fgColor)

    def test_set_patter_fill_raise_exception(self):
        with self.assertRaises(expected_exception=ValueError):
            self.excel_repository.set_pattern_fill(start_row=1, start_column=1, end_row=1, end_column=1,
                                                   pattern_type='aaa',
                                                   fg_color='d3d3d3')

    def test_set_data_validation(self):
        formula = ['red', 'blue', 'yellow']
        self.excel_repository.set_data_validation(start_row=1, end_row=31, column=1, formula=formula)

    def test_set_data_validation_from_cell_range(self):
        cell_range = ExcelUtil.create_cell_range_string(start_column='C', start_row=5, end_column='C', end_row=10,
                                                        fixed_start_column=True, fixed_start_row=True,
                                                        fixed_end_column=True, fixed_end_row=True)
        self.excel_repository.set_data_validation_from_cell_range(start_row=1, end_row=31, column=1,
                                                                  cell_range=cell_range)

    def test_set_number_format(self):
        self.excel_repository.set_number_format(row=1, column=1, number_format='[h]:mm')
