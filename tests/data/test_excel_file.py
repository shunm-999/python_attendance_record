import os
import unittest

import openpyxl

from data.excel_file import ExcelFile


class TestCreateExcelFile(unittest.TestCase):

    def setUp(self) -> None:
        self.excel_filename = 'sample_data.xlsx'
        openpyxl.Workbook().save(self.excel_filename)

    def tearDown(self) -> None:
        try:
            os.remove(self.excel_filename)
        except IOError:
            pass

    def test_load_excel_file_raise_exception(self):
        with self.assertRaises(expected_exception=FileNotFoundError):
            ExcelFile(filename="raise_exception.xlsx")

    def test_load_excel_file(self):
        ExcelFile(filename=self.excel_filename)

    def test_create_excel_file(self):
        excel_filename = 'create.xlsx'
        ExcelFile(filename=excel_filename, create_file=True)


class TestOperationExcelFile(unittest.TestCase):
    def setUp(self) -> None:
        self.excel_filename = 'sample_data.xlsx'
        self.excel_file = ExcelFile(filename=self.excel_filename, create_file=True)
        self.excel_file.delete_sheet_from_index(0)

    def tearDown(self) -> None:
        try:
            os.remove(self.excel_filename)
        except IOError:
            pass

    def test_is_contains(self):
        self.excel_file.create_sheet('sample')
        self.assertTrue(self.excel_file.is_contains('sample'))

    def test_get_sheet_from_index(self):
        expected = 'sample'
        self.excel_file.create_sheet(expected)

        actual = self.excel_file.get_sheet_from_index(0).get_title()
        self.assertEqual(expected, actual)

    def test_get_sheet_from_index_none(self):
        self.assertIsNone(self.excel_file.get_sheet_from_index(0))

    def test_get_sheet_from_sheet_title(self):
        expected = 'sample'
        self.excel_file.create_sheet(expected)

        actual = self.excel_file.get_sheet_from_sheet_title(expected).get_title()
        self.assertEqual(expected, actual)

    def test_get_sheet_from_sheet_title_none(self):
        self.assertIsNone(self.excel_file.get_sheet_from_sheet_title('sample'))

    def test_get_active_sheet(self):
        expected = 'sample'
        self.excel_file.create_sheet(expected)

        actual = self.excel_file.active_sheet.get_title()
        self.assertEqual(expected, actual)

    def test_get_active_sheet_none(self):
        self.assertIsNone(self.excel_file.active_sheet)

    def test_create_sheet(self):
        expected_sheet1 = 'sample'
        expected_sheet2 = 'sample2'

        self.excel_file.create_sheet(expected_sheet1)
        actual_sheet1 = self.excel_file.get_sheet_from_sheet_title(expected_sheet1).get_title()
        self.assertEqual(expected_sheet1, actual_sheet1)

        self.excel_file.create_sheet(expected_sheet2)
        actual_sheet2 = self.excel_file.get_sheet_from_sheet_title(expected_sheet2).get_title()
        self.assertEqual(expected_sheet2, actual_sheet2)

    def test_insert_sheet(self):
        expected_sheet1 = 'sample'
        expected_sheet2 = 'sample2'

        self.excel_file.create_sheet(expected_sheet1, index=0)
        actual_sheet1 = self.excel_file.get_sheet_from_index(0).get_title()
        self.assertEqual(expected_sheet1, actual_sheet1)

        self.excel_file.create_sheet(expected_sheet2, index=0)
        actual_sheet2 = self.excel_file.get_sheet_from_index(0).get_title()
        self.assertEqual(expected_sheet2, actual_sheet2)

    def test_delete_sheet_from_index(self):
        self.excel_file.create_sheet('sample')
        self.assertEqual(1, self.excel_file.get_sheet_count())

        self.excel_file.delete_sheet_from_index(0)
        self.assertEqual(0, self.excel_file.get_sheet_count())

    def test_delete_sheet_from_wrong_index(self):
        self.excel_file.create_sheet('sample')
        self.assertEqual(1, self.excel_file.get_sheet_count())

        self.excel_file.delete_sheet_from_index(2)
        self.assertEqual(1, self.excel_file.get_sheet_count())

    def test_delete_current_sheet_from_index(self):

        expected_sheet1 = 'sample'
        expected_sheet2 = 'sample2'

        self.excel_file.create_sheet(expected_sheet1)
        self.excel_file.create_sheet(expected_sheet2)

        self.assertEqual(expected_sheet2, self.excel_file.active_sheet.get_title())

        self.excel_file.delete_sheet_from_index(1)
        self.assertEqual(1, self.excel_file.get_sheet_count())
        self.assertIsNone(self.excel_file.active_sheet)

        self.excel_file.switch_active_sheet_from_index(0)
        self.assertEqual(expected_sheet1, self.excel_file.active_sheet.get_title())

    def test_delete_sheet_from_sheet_title(self):
        self.excel_file.create_sheet('sample')
        self.assertEqual(1, self.excel_file.get_sheet_count())

        self.excel_file.delete_sheet_from_sheet_title('sample')
        self.assertEqual(0, self.excel_file.get_sheet_count())

    def test_delete_sheet_from_wrong_sheet_title(self):
        self.excel_file.create_sheet('sample')
        self.assertEqual(1, self.excel_file.get_sheet_count())

        self.excel_file.delete_sheet_from_sheet_title('wrong')
        self.assertEqual(1, self.excel_file.get_sheet_count())

    def test_delete_current_sheet_from_title(self):

        expected_sheet1 = 'sample'
        expected_sheet2 = 'sample2'

        self.excel_file.create_sheet(expected_sheet1)
        self.excel_file.create_sheet(expected_sheet2)

        self.assertEqual(expected_sheet2, self.excel_file.active_sheet.get_title())

        self.excel_file.delete_sheet_from_sheet_title(expected_sheet2)
        self.assertEqual(1, self.excel_file.get_sheet_count())
        self.assertIsNone(self.excel_file.active_sheet)

        self.excel_file.switch_active_sheet_from_sheet_title(expected_sheet1)
        self.assertEqual(expected_sheet1, self.excel_file.active_sheet.get_title())

    def test_get_sheet_count(self):
        self.excel_file.create_sheet('sample')
        self.excel_file.create_sheet('sample2')

        self.assertEqual(2, self.excel_file.get_sheet_count())

    def test_switch_active_sheet_from_index(self):

        expected1 = 'sample'
        expected2 = 'sample2'

        self.excel_file.create_sheet(expected1)
        self.excel_file.create_sheet(expected2)
        excel_sheet = self.excel_file.active_sheet
        self.assertEqual(expected2, excel_sheet.get_title())

        self.excel_file.switch_active_sheet_from_index(0)
        excel_sheet = self.excel_file.active_sheet
        self.assertEqual(expected1, excel_sheet.get_title())

    def test_switch_active_sheet_from_wrong_index(self):

        expected1 = 'sample'
        expected2 = 'sample2'

        self.excel_file.create_sheet(expected1)
        self.excel_file.create_sheet(expected2)
        excel_sheet = self.excel_file.active_sheet
        self.assertEqual(expected2, excel_sheet.get_title())

        self.excel_file.switch_active_sheet_from_index(2)
        excel_sheet = self.excel_file.active_sheet
        self.assertEqual(expected2, excel_sheet.get_title())

    def test_switch_active_sheet_from_sheet_title(self):

        expected1 = 'sample'
        expected2 = 'sample2'

        self.excel_file.create_sheet(expected1)
        self.excel_file.create_sheet(expected2)
        excel_sheet = self.excel_file.active_sheet
        self.assertEqual(expected2, excel_sheet.get_title())

        self.excel_file.switch_active_sheet_from_sheet_title(expected1)
        excel_sheet = self.excel_file.active_sheet
        self.assertEqual(expected1, excel_sheet.get_title())

    def test_switch_active_sheet_from_wrong_sheet_title(self):

        expected1 = 'sample'
        expected2 = 'sample2'

        self.excel_file.create_sheet(expected1)
        self.excel_file.create_sheet(expected2)
        excel_sheet = self.excel_file.active_sheet
        self.assertEqual(expected2, excel_sheet.get_title())

        self.excel_file.switch_active_sheet_from_sheet_title('wrong')
        excel_sheet = self.excel_file.active_sheet
        self.assertEqual(expected2, excel_sheet.get_title())

    def test_save(self):
        self.excel_file.create_sheet('sample')
        self.assertFalse(os.path.exists(self.excel_filename))

        self.excel_file.save()
        self.assertTrue(os.path.exists(self.excel_filename))

    def test_save_new_title(self):

        self.excel_file.create_sheet('sample')
        self.assertFalse(os.path.exists(self.excel_filename))

        self.excel_file.save(filename=self.excel_filename)
        self.assertTrue(os.path.exists(self.excel_filename))

    def test_delete(self):
        self.excel_file.create_sheet('sample')
        self.excel_file.save()

        self.assertTrue(os.path.exists(self.excel_filename))

        self.excel_file.delete()
        self.assertFalse(os.path.exists(self.excel_filename))

    def test_delete_not_exist_file(self):
        self.excel_file.delete()
