import unittest

import openpyxl
from openpyxl.styles import Color
from openpyxl.styles.colors import BLACK

from dao.excel_sheet_dao import ExcelSheetDao
from data.excel_sheet import ExcelSheet


class TestExcelDao(unittest.TestCase):
    def setUp(self) -> None:
        workbook = openpyxl.Workbook()
        excel_sheet = ExcelSheet(worksheet=workbook.active)
        self.excel_sheet_dao = ExcelSheetDao(excel_sheet)

    def test_get_empty_cell(self):
        expected = None
        actual = self.excel_sheet_dao.get(row=1, column=1)
        self.assertEqual(expected, actual)

    def test_get_cell_from_column_number(self):
        expected = 1
        self.excel_sheet_dao.put(row=1, column=1, value=expected)

        actual = self.excel_sheet_dao.get(row=1, column=1)
        self.assertEqual(expected, actual)

    def test_get_cell_from_column_letter(self):
        expected = 1
        self.excel_sheet_dao.put(row=1, column='A', value=expected)

        actual = self.excel_sheet_dao.get(row=1, column='A')
        self.assertEqual(expected, actual)

    def test_set_and_get_rows(self):
        expected = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]

        for row_num in range(1, 4):
            for column_num in range(1, 4):
                self.excel_sheet_dao.put(row=row_num, column=column_num, value=expected[row_num - 1][column_num - 1])

        actual = self.excel_sheet_dao.get_rows(start_row=1, start_column=1, end_row=3, end_column=3)
        self.assertEqual(expected, actual)

    def test_insert_rows(self):
        expected = 1
        self.excel_sheet_dao.put(row=1, column=1, value=expected)
        self.excel_sheet_dao.insert_rows(row=1)

        actual = self.excel_sheet_dao.get(row=2, column=1)
        self.assertEqual(expected, actual)

    def test_delete_rows(self):
        expected = 1
        self.excel_sheet_dao.put(row=2, column=1, value=expected)
        self.excel_sheet_dao.delete_rows(row=1)

        actual = self.excel_sheet_dao.get(row=1, column=1)
        self.assertEqual(expected, actual)

    def test_insert_columns(self):
        expected = 1
        self.excel_sheet_dao.put(row=1, column=1, value=expected)
        self.excel_sheet_dao.insert_columns(column=1)

        actual = self.excel_sheet_dao.get(row=1, column=2)
        self.assertEqual(expected, actual)

    def test_delete_columns(self):
        expected = 1
        self.excel_sheet_dao.put(row=1, column=2, value=expected)
        self.excel_sheet_dao.delete_columns(column=1)

        actual = self.excel_sheet_dao.get(row=1, column=1)
        self.assertEqual(expected, actual)

    def test_set_and_get_row_height(self):
        expected = 16
        self.excel_sheet_dao.set_row_height(row=1, height=expected)
        actual = self.excel_sheet_dao.get_row_height(row=1)

        self.assertEqual(expected, actual)

    def test_set_and_get_column_width(self):
        expected = 16
        self.excel_sheet_dao.set_column_width(column=1, width=expected)
        actual = self.excel_sheet_dao.get_column_width(column=1)

        self.assertEqual(expected, actual)

    def test_merge_cells(self):

        self.excel_sheet_dao.put(row=1, column=1, value=1)
        self.excel_sheet_dao.put(row=3, column=3, value=3)
        self.excel_sheet_dao.merge_cells(start_row=1, start_column=1, end_row=3, end_column=3)

        expected_a1 = 1

        actual_a1 = self.excel_sheet_dao.get(row=1, column=1)
        actual_c3 = self.excel_sheet_dao.get(row=3, column=3)

        self.assertEqual(expected_a1, actual_a1)
        self.assertIsNone(actual_c3)

    def test_merge_cells_and_dimensions(self):

        for row_num in range(1, 4):
            self.excel_sheet_dao.set_row_height(row=row_num, height=16)
        for column_num in range(1, 4):
            self.excel_sheet_dao.set_column_width(column=column_num, width=16)

        self.excel_sheet_dao.merge_cells_and_dimensions(start_row=1, start_column=1, end_row=3, end_column=3)

        expected = 48

        actual_height = self.excel_sheet_dao.get_row_height(row=1)
        actual_width = self.excel_sheet_dao.get_column_width(column=1)

        self.assertEqual(expected, actual_height)
        self.assertEqual(expected, actual_width)

    def test_freeze_panes(self):
        self.excel_sheet_dao.set_freeze_panes(row=2, column=1)

    def test_freeze_panes_column_letter(self):
        self.excel_sheet_dao.set_freeze_panes(row=2, column='B')

    def test_set_and_get_font(self):
        self.excel_sheet_dao.set_font(row=1, column=1, size=16, bold=True, italic=True)
        font = self.excel_sheet_dao.get_font(row=1, column=1)

        self.assertEqual(font.size, 16)
        self.assertTrue(font.bold)
        self.assertTrue(font.italic)

    def test_set_and_get_font_column_letter(self):
        self.excel_sheet_dao.set_font(row=1, column='A', size=16, bold=True, italic=True)
        font = self.excel_sheet_dao.get_font(row=1, column='A')

        self.assertEqual(font.size, 16)
        self.assertTrue(font.bold)
        self.assertTrue(font.italic)

    def test_set_ruled_line(self):
        self.excel_sheet_dao.set_ruled_line(start_row=1, start_column=1, end_row=3, end_column=3)

        for row_num in range(1, 4):
            for column_num in range(1, 4):
                border = self.excel_sheet_dao.get_border(row=row_num, column=column_num)
                self.assertIsNotNone(border)

    def test_set_ruled_line_raise_exception(self):
        with self.assertRaises(expected_exception=ValueError):
            self.excel_sheet_dao.set_ruled_line(start_row=1, start_column=1, end_row=3, end_column=3, style='aaa')

    def test_set_and_get_pattern_fill(self):

        expected_pattern_type = 'solid'
        expected_fg_color = '00d3d3d3'

        self.excel_sheet_dao.set_pattern_fill(start_row=1, start_column=1, end_row=1, end_column=1,
                                              pattern_type=expected_pattern_type,
                                              fg_color=expected_fg_color)

        pattern_fill = self.excel_sheet_dao.get_pattern_fill(row=1, column=1)
        self.assertEqual(expected_pattern_type, pattern_fill.patternType)
        self.assertEqual(expected_fg_color, pattern_fill.fgColor.value)

    def test_set_and_get_color_from_name(self):

        expected_fg_color = Color(rgb=BLACK)
        self.excel_sheet_dao.set_pattern_fill(start_row=1, start_column=1, end_row=1, end_column=1,
                                              pattern_type='solid',
                                              fg_color=expected_fg_color)

        pattern_fill = self.excel_sheet_dao.get_pattern_fill(row=1, column=1)
        self.assertEqual(expected_fg_color, pattern_fill.fgColor)

    def test_set_patter_fill_raise_exception(self):
        with self.assertRaises(expected_exception=ValueError):
            self.excel_sheet_dao.set_pattern_fill(start_row=1, start_column=1, end_row=1, end_column=1,
                                                  pattern_type='aaa',
                                                  fg_color='d3d3d3')

    def test_set_data_validation(self):
        formula = ['red', 'blue', 'yellow']
        self.excel_sheet_dao.set_data_validation(start_row=1, end_row=31, column=1, formula=formula)
