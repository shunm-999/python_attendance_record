from openpyxl.styles import Font, PatternFill, Border

from const.alignmentconst import AlignmentConst
from data.excel_sheet import ExcelSheet
from util.excel_util import ExcelUtil


class ExcelSheetDao:
    """
    基本関数
    """

    def __init__(self, excel_sheet: ExcelSheet):
        self.__excel_sheet = excel_sheet

    def get(self, row, column):
        return self.__excel_sheet.get(row=row, column=column)

    def get_rows(self, start_row, start_column, end_row=None, end_column=None):
        return self.__excel_sheet.get_rows(start_row=start_row, start_column=start_column, end_row=end_row,
                                           end_column=end_column)

    def put(self, row, column, value):
        self.__excel_sheet.put(row=row, column=column, value=value)

    def put_rows(self, start_row, start_column, end_row=None, end_column=None, values=None):
        self.__excel_sheet.put_rows(start_row=start_row,
                                    start_column=start_column,
                                    end_row=end_row,
                                    end_column=end_column,
                                    values=values)

    def insert_rows(self, row):
        self.__excel_sheet.insert_rows(row=row)

    def delete_rows(self, row):
        self.__excel_sheet.delete_rows(row=row)

    def insert_columns(self, column):
        self.__excel_sheet.insert_columns(column=column)

    def delete_columns(self, column):
        self.__excel_sheet.delete_columns(column=column)

    def get_row_height(self, row) -> int:
        return self.__excel_sheet.get_row_height(row=row)

    def set_row_height(self, row, height):
        self.__excel_sheet.set_row_height(row=row, height=height)

    def get_column_width(self, column) -> int:
        return self.__excel_sheet.get_column_width(column=column)

    def set_column_width(self, column, width):
        self.__excel_sheet.set_column_width(column=column, width=width)

    def merge_cells(self, start_row, start_column, end_row, end_column):
        self.__excel_sheet.merge_cells(start_row=start_row,
                                       start_column=start_column,
                                       end_row=end_row,
                                       end_column=end_column)

    def set_freeze_panes(self, row, column):
        self.__excel_sheet.set_freeze_panes(row=row, column=column)

    def get_font(self, row, column) -> Font:
        return self.__excel_sheet.get_font(row=row, column=column)

    def set_font(self, row, column, size=None, bold=None, italic=None):
        self.__excel_sheet.set_font(row=row, column=column, size=size, bold=bold, italic=italic)

    def get_border(self, row, column) -> Border:
        return self.__excel_sheet.get_border(row=row, column=column)

    def set_border(self, row, column, top=False, bottom=False, left=False, right=False, style='thin',
                   color='000000') -> None:
        self.__excel_sheet.set_border(row=row,
                                      column=column,
                                      top=top,
                                      bottom=bottom,
                                      left=left,
                                      right=right,
                                      style=style,
                                      color=color)

    def set_ruled_line(self, start_row, start_column, end_row=None, end_column=None, style='thin', color='000000'):
        self.__excel_sheet.set_ruled_line(start_row=start_row,
                                          start_column=start_column,
                                          end_row=end_row,
                                          end_column=end_column,
                                          style=style,
                                          color=color)

    def get_pattern_fill(self, row, column) -> PatternFill:
        return self.__excel_sheet.get_pattern_fill(row=row, column=column)

    def set_pattern_fill(self,
                         start_row,
                         start_column,
                         end_row=None,
                         end_column=None,
                         pattern_type='solid',
                         fg_color='d3d3d3'):
        self.__excel_sheet.set_pattern_fill(start_row=start_row,
                                            start_column=start_column,
                                            end_row=end_row,
                                            end_column=end_column,
                                            pattern_type=pattern_type,
                                            fg_color=fg_color)

    def set_data_validation(self, start_row, end_row, column, formula=None):
        self.__excel_sheet.set_data_validation(start_row=start_row, end_row=end_row, column=column, formula=formula)

    def set_data_validation_from_cell_range(self, start_row, end_row, column, cell_range: str):
        self.__excel_sheet.set_data_validation_from_cell_range(start_row=start_row,
                                                               end_row=end_row,
                                                               column=column,
                                                               cell_range=cell_range)

    def set_number_format(self, row, column, number_format: str):
        self.__excel_sheet.set_number_format(row=row, column=column, number_format=number_format)

    def set_number_formats(self, start_row, start_column, end_row, end_column, number_format: str):
        self.__excel_sheet.set_number_formats(start_row=start_row, start_column=start_column, end_row=end_row,
                                              end_column=end_column, number_format=number_format)

    def set_text_alignment(self, row, column, horizontal: AlignmentConst, vertical: AlignmentConst,
                           text_rotation: int = 0):
        self.__excel_sheet.set_text_alignment(row=row, column=column, horizontal=horizontal, vertical=vertical,
                                              text_rotation=text_rotation)

    def set_text_alignments(self, start_row, start_column, end_row, end_column, horizontal: AlignmentConst,
                            vertical: AlignmentConst, text_rotation: int = 0):
        self.__excel_sheet.set_text_alignments(start_row=start_row, start_column=start_column, end_row=end_row,
                                               end_column=end_column, horizontal=horizontal,
                                               vertical=vertical, text_rotation=text_rotation)

    """
    Excel関数
    """

    def put_sum(self, target_row, target_column, start_row, start_column, end_row, end_column):
        cell_range = ExcelUtil.create_cell_range_string(start_row=start_row,
                                                        start_column=start_column,
                                                        end_row=end_row,
                                                        end_column=end_column)

        self.__excel_sheet.put(row=target_row, column=target_column, value=f'=SUM({cell_range})')

    def put_average(self, target_row, target_column, start_row, start_column, end_row, end_column):
        cell_range = ExcelUtil.create_cell_range_string(start_row=start_row,
                                                        start_column=start_column,
                                                        end_row=end_row,
                                                        end_column=end_column)

        self.__excel_sheet.put(row=target_row, column=target_column, value=f'=AVERAGE({cell_range})')
