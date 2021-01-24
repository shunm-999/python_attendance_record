from openpyxl.utils import get_column_letter


class ExcelUtil:

    @staticmethod
    def create_cell_string(row, column, fixed_row=False, fixed_column=False):
        if type(column) is int:
            column = get_column_letter(column)

        row = f'${row}' if fixed_row else row
        column = f'${column}' if fixed_column else column

        return f'{column}{row}'

    @staticmethod
    def create_cell_range_string(start_row,
                                 start_column,
                                 end_row,
                                 end_column,
                                 fixed_start_row=False,
                                 fixed_start_column=False,
                                 fixed_end_row=False,
                                 fixed_end_column=False):
        if type(start_column) is int:
            start_column = get_column_letter(start_column)

        if type(end_column) is int:
            end_column = get_column_letter(end_column)

        start_row = f'${start_row}' if fixed_start_row else start_row
        start_column = f'${start_column}' if fixed_start_column else start_column
        end_row = f'${end_row}' if fixed_end_row else end_row
        end_column = f'${end_column}' if fixed_end_column else end_column

        return f'{start_column}{start_row}:{end_column}{end_row}'
